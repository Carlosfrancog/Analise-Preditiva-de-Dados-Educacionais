#!/usr/bin/env python3
"""
Sistema de Cadastro Escolar - Alunos, Matérias e Notas
Gera planilha Excel formatada com médias dos alunos
"""

import sqlite3
import random
import sys
from pathlib import Path

DB_PATH = "escola.db"

SALAS = {
    "6º Fundamental": "6F",
    "7º Fundamental": "7F",
    "8º Fundamental": "8F",
    "9º Fundamental": "9F",
    "1º Médio": "1M",
    "2º Médio": "2M",
    "3º Médio": "3M",
}

NOMES = [
    "Ana","Bruno","Carla","Carlos","Daniel","Daniela","Eduardo","Fernanda",
    "Gabriel","Helena","Igor","João","Juliana","Lucas","Mariana","Pedro",
    "Rafael","Sofia","Tiago","Vanessa","Beatriz","Diego","Leticia","Mateus",
    "Natalia","Paulo","Renata","Sergio","Tatiane","Victor","Yasmin","Leandro",
    "Camila","Felipe","Gabriela","Hugo","Isabela","Jorge","Karina","Leonardo",
]
SOBRENOMES = [
    "Lima","Silva","Santos","Oliveira","Souza","Costa","Ferreira","Pereira",
    "Rodrigues","Alves","Nascimento","Carvalho","Gomes","Martins","Araújo",
    "Ribeiro","Azevedo","Cardoso","Rocha","Moreira","Nunes","Barbosa",
]

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS salas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL,
            codigo TEXT UNIQUE NOT NULL
        );
        CREATE TABLE IF NOT EXISTS alunos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            sala_id INTEGER NOT NULL,
            matricula TEXT UNIQUE NOT NULL,
            FOREIGN KEY (sala_id) REFERENCES salas(id)
        );
        CREATE TABLE IF NOT EXISTS materias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        );
        CREATE TABLE IF NOT EXISTS notas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id INTEGER NOT NULL,
            materia_id INTEGER NOT NULL,
            n1 REAL,
            n2 REAL,
            n3 REAL,
            n4 REAL,
            UNIQUE(aluno_id, materia_id),
            FOREIGN KEY (aluno_id) REFERENCES alunos(id),
            FOREIGN KEY (materia_id) REFERENCES materias(id)
        );

        CREATE TABLE IF NOT EXISTS ml_features (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id            INTEGER NOT NULL,
            materia_id          INTEGER NOT NULL,
            aluno_nome          TEXT,
            materia_nome        TEXT,
            sala_nome           TEXT,
            serie_num           INTEGER,
            n1                  REAL,
            n2                  REAL,
            n3                  REAL,
            n4                  REAL,
            media_ponderada     REAL,
            n1_norm             REAL,
            n2_norm             REAL,
            n3_norm             REAL,
            n4_norm             REAL,
            media_pond_norm     REAL,
            media_geral_aluno   REAL,
            slope_notas         REAL,
            variancia_notas     REAL,
            serie_num_norm      REAL,
            pct_materias_ok     REAL,
            media_turma_norm    REAL,
            status_encoded      INTEGER,
            status_label        TEXT,
            gerado_em           TEXT DEFAULT (datetime('now')),
            UNIQUE(aluno_id, materia_id),
            FOREIGN KEY (aluno_id)   REFERENCES alunos(id),
            FOREIGN KEY (materia_id) REFERENCES materias(id)
        );
    """)
    # Insert default salas
    for nome, codigo in SALAS.items():
        c.execute("INSERT OR IGNORE INTO salas (nome, codigo) VALUES (?,?)", (nome, codigo))
    conn.commit()
    conn.close()
    _migrate_db()   # aplica migrações em bancos antigos


def _migrate_db():
    """
    Garante que todas as colunas e tabelas novas existam em bancos antigos.
    Seguro rodar múltiplas vezes — usa IF NOT EXISTS / ignora erros de coluna já existente.
    """
    conn = get_conn()
    c = conn.cursor()

    # Cria ml_features se não existir (banco criado antes desta versão)
    c.executescript("""
        CREATE TABLE IF NOT EXISTS ml_features (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id            INTEGER NOT NULL,
            materia_id          INTEGER NOT NULL,
            aluno_nome          TEXT,
            materia_nome        TEXT,
            sala_nome           TEXT,
            serie_num           INTEGER,
            n1                  REAL,
            n2                  REAL,
            n3                  REAL,
            n4                  REAL,
            media_ponderada     REAL,
            n1_norm             REAL,
            n2_norm             REAL,
            n3_norm             REAL,
            n4_norm             REAL,
            media_pond_norm     REAL,
            media_geral_aluno   REAL,
            slope_notas         REAL,
            variancia_notas     REAL,
            serie_num_norm      REAL,
            pct_materias_ok     REAL,
            media_turma_norm    REAL,
            status_encoded      INTEGER,
            status_label        TEXT,
            gerado_em           TEXT DEFAULT (datetime('now')),
            UNIQUE(aluno_id, materia_id),
            FOREIGN KEY (aluno_id)   REFERENCES alunos(id),
            FOREIGN KEY (materia_id) REFERENCES materias(id)
        );
    """)

    # Adiciona colunas novas que possam faltar em versões antigas da tabela
    existing_cols = {
        row[1] for row in c.execute("PRAGMA table_info(ml_features)").fetchall()
    }
    migrations = [
        ("gerado_em",         "TEXT DEFAULT (datetime('now'))"),
        ("media_ponderada",   "REAL"),
        ("media_geral_aluno", "REAL"),
        ("slope_notas",       "REAL"),
        ("variancia_notas",   "REAL"),
        ("serie_num_norm",    "REAL"),
        ("pct_materias_ok",   "REAL"),
        ("media_turma_norm",  "REAL"),
    ]
    for col, typedef in migrations:
        if col not in existing_cols:
            try:
                c.execute(f"ALTER TABLE ml_features ADD COLUMN {col} {typedef}")
            except Exception:
                pass  # já existe ou erro irrelevante

    conn.commit()
    conn.close()

def gerar_matricula(sala_id):
    conn = get_conn()
    count = conn.execute("SELECT COUNT(*) FROM alunos WHERE sala_id=?", (sala_id,)).fetchone()[0]
    sala = conn.execute("SELECT codigo FROM salas WHERE id=?", (sala_id,)).fetchone()
    conn.close()
    return f"{sala['codigo']}{count+1:04d}"

def adicionar_aluno(nome, sala_id):
    conn = get_conn()
    matricula = gerar_matricula(sala_id)
    conn.execute("INSERT INTO alunos (nome, sala_id, matricula) VALUES (?,?,?)", (nome, sala_id, matricula))
    conn.commit()
    conn.close()

def gerar_alunos_genericos(quantidade, sala_id):
    conn = get_conn()
    usados = set(r[0] for r in conn.execute("SELECT nome FROM alunos WHERE sala_id=?", (sala_id,)).fetchall())
    adicionados = 0
    tentativas = 0
    while adicionados < quantidade and tentativas < quantidade * 5:
        nome = f"{random.choice(NOMES)} {random.choice(SOBRENOMES)}"
        tentativas += 1
        if nome not in usados:
            usados.add(nome)
            sala = conn.execute("SELECT codigo FROM salas WHERE id=?", (sala_id,)).fetchone()
            count = conn.execute("SELECT COUNT(*) FROM alunos WHERE sala_id=?", (sala_id,)).fetchone()[0]
            matricula = f"{sala['codigo']}{count+1:04d}"
            conn.execute("INSERT INTO alunos (nome, sala_id, matricula) VALUES (?,?,?)", (nome, sala_id, matricula))
            conn.commit()
            adicionados += 1
    conn.close()
    return adicionados

def adicionar_materia(nome):
    conn = get_conn()
    conn.execute("INSERT OR IGNORE INTO materias (nome) VALUES (?)", (nome,))
    conn.commit()
    conn.close()

def atribuir_materias_todos():
    conn = get_conn()
    alunos = conn.execute("SELECT id FROM alunos").fetchall()
    materias = conn.execute("SELECT id FROM materias").fetchall()
    count = 0
    for a in alunos:
        for m in materias:
            try:
                conn.execute("INSERT OR IGNORE INTO notas (aluno_id, materia_id) VALUES (?,?)", (a['id'], m['id']))
                count += 1
            except:
                pass
    conn.commit()
    conn.close()
    return count

def gerar_notas_aleatorias(aluno_ids=None, materia_ids=None):
    conn = get_conn()
    query = "SELECT id FROM notas WHERE 1=1"
    params = []
    if aluno_ids:
        query += f" AND aluno_id IN ({','.join('?'*len(aluno_ids))})"
        params.extend(aluno_ids)
    if materia_ids:
        query += f" AND materia_id IN ({','.join('?'*len(materia_ids))})"
        params.extend(materia_ids)
    rows = conn.execute(query, params).fetchall()
    for r in rows:
        n1 = round(random.uniform(2, 10), 1)
        n2 = round(random.uniform(2, 10), 1)
        n3 = round(random.uniform(2, 10), 1)
        n4 = round(random.uniform(2, 10), 1)
        conn.execute("UPDATE notas SET n1=?,n2=?,n3=?,n4=? WHERE id=?", (n1, n2, n3, n4, r['id']))
    conn.commit()
    conn.close()
    return len(rows)

def salvar_nota(aluno_id, materia_id, n1, n2, n3, n4):
    conn = get_conn()
    conn.execute("""
        INSERT INTO notas (aluno_id, materia_id, n1, n2, n3, n4)
        VALUES (?,?,?,?,?,?)
        ON CONFLICT(aluno_id, materia_id)
        DO UPDATE SET n1=excluded.n1, n2=excluded.n2, n3=excluded.n3, n4=excluded.n4
    """, (aluno_id, materia_id, n1, n2, n3, n4))
    conn.commit()
    conn.close()

def get_salas():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM salas ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_alunos(sala_id=None):
    conn = get_conn()
    if sala_id:
        rows = conn.execute("""
            SELECT a.*, s.nome as sala_nome, s.codigo as sala_codigo
            FROM alunos a JOIN salas s ON a.sala_id=s.id
            WHERE a.sala_id=? ORDER BY a.nome
        """, (sala_id,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT a.*, s.nome as sala_nome, s.codigo as sala_codigo
            FROM alunos a JOIN salas s ON a.sala_id=s.id
            ORDER BY s.id, a.nome
        """).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_materias():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM materias ORDER BY nome").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_notas(aluno_id):
    conn = get_conn()
    rows = conn.execute("""
        SELECT n.*, m.nome as materia_nome
        FROM notas n JOIN materias m ON n.materia_id=m.id
        WHERE n.aluno_id=?
        ORDER BY m.nome
    """, (aluno_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_relatorio(sala_id=None):
    conn = get_conn()
    query = """
        SELECT a.nome as aluno, s.nome as sala, m.nome as materia,
               n.n1, n.n2, n.n3, n.n4
        FROM notas n
        JOIN alunos a ON n.aluno_id=a.id
        JOIN salas s ON a.sala_id=s.id
        JOIN materias m ON n.materia_id=m.id
        WHERE (n.n1 IS NOT NULL OR n.n2 IS NOT NULL OR n.n3 IS NOT NULL OR n.n4 IS NOT NULL)
    """
    params = []
    if sala_id:
        query += " AND a.sala_id=?"
        params.append(sala_id)
    query += " ORDER BY s.id, a.nome, m.nome"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Pesos de cada avaliação na média ponderada ────────────────────────────────
# N1: início do bimestre (peso menor), N4: avaliação final (peso maior)
PESOS_NOTAS = {"n1": 0.20, "n2": 0.25, "n3": 0.25, "n4": 0.30}

# Mapeamento de série → número ordinal para normalização
SERIE_MAP = {
    "6F": 6, "7F": 7, "8F": 8, "9F": 9,
    "1M": 10, "2M": 11, "3M": 12,
}
SERIE_MIN, SERIE_MAX = 6, 12


def _slope(vals):
    """Tendência linear simples (n1→n4). Retorna valor em -1..+1."""
    xs = [i for i, v in enumerate(vals) if v is not None]
    ys = [v for v in vals if v is not None]
    if len(xs) < 2:
        return 0.0
    n = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    den = sum((x - mx) ** 2 for x in xs)
    raw = (num / den) if den else 0.0
    # Normaliza: max slope possível é 10/3 (notas 0→10 em 3 passos)
    return max(-1.0, min(1.0, raw / (10 / max(len(xs) - 1, 1))))


def _std(vals):
    """Desvio padrão das notas, normalizado para 0–1 (max std = 5)."""
    clean = [v for v in vals if v is not None]
    if len(clean) < 2:
        return 0.0
    m = sum(clean) / len(clean)
    variance = sum((v - m) ** 2 for v in clean) / len(clean)
    return min(1.0, (variance ** 0.5) / 5.0)


def gerar_features_ml(sala_id=None):
    """
    Calcula e persiste features de ML na tabela ml_features.

    Para cada par (aluno, materia) com pelo menos uma nota preenchida:
      - Calcula média ponderada (pesos n1=0.20, n2=0.25, n3=0.25, n4=0.30)
      - Normaliza todas as features para 0–1
      - Determina label: 0=Reprovado, 1=Recuperação, 2=Aprovado
      - Salva em ml_features (upsert)

    Retorna: (total_gerado, dict com estatísticas resumidas)
    """
    conn = get_conn()

    # ── Busca todos os registros de notas com metadados ──────────────────────
    q = """
        SELECT
            n.aluno_id, n.materia_id,
            a.nome  AS aluno_nome,
            m.nome  AS materia_nome,
            s.nome  AS sala_nome,
            s.codigo AS sala_codigo,
            n.n1, n.n2, n.n3, n.n4
        FROM notas n
        JOIN alunos   a ON n.aluno_id   = a.id
        JOIN materias m ON n.materia_id = m.id
        JOIN salas    s ON a.sala_id    = s.id
    """
    if sala_id:
        q += " WHERE a.sala_id = ?"
        rows = conn.execute(q, (sala_id,)).fetchall()
    else:
        rows = conn.execute(q).fetchall()

    if not rows:
        conn.close()
        return 0, {}

    # ── Pré-calcula agregados por aluno e por (turma × matéria) ─────────────
    # média geral de cada aluno (todas as matérias)
    medias_aluno = {}      # aluno_id → [média_pond, ...]
    # média da turma por matéria
    medias_turma = {}      # (sala_codigo, materia_id) → [média_pond, ...]
    # qtd matérias com média >= 6 por aluno
    ok_aluno = {}          # aluno_id → (aprovadas, total)

    def media_pond(n1, n2, n3, n4):
        notas = [n1, n2, n3, n4]
        pesos = [0.20, 0.25, 0.25, 0.30]
        soma = sum(p * v for p, v in zip(pesos, notas) if v is not None)
        sp   = sum(p     for p, v in zip(pesos, notas) if v is not None)
        return round(soma / sp, 4) if sp > 0 else None

    # Primeiro passo: calcular médias para os agregados
    temp = []
    for r in rows:
        mp = media_pond(r["n1"], r["n2"], r["n3"], r["n4"])
        temp.append((r, mp))
        if mp is None:
            continue
        medias_aluno.setdefault(r["aluno_id"], []).append(mp)
        medias_turma.setdefault((r["sala_codigo"], r["materia_id"]), []).append(mp)

    for r, mp in temp:
        if mp is None:
            continue
        ok = ok_aluno.get(r["aluno_id"], [0, 0])
        ok[1] += 1
        if mp >= 6.0:
            ok[0] += 1
        ok_aluno[r["aluno_id"]] = ok

    # ── Gera e persiste features ─────────────────────────────────────────────
    total = 0
    stats = {"aprovado": 0, "recuperacao": 0, "reprovado": 0, "sem_nota": 0}

    for r, mp in temp:
        n1, n2, n3, n4 = r["n1"], r["n2"], r["n3"], r["n4"]
        notas = [n1, n2, n3, n4]

        # Pula registros sem nenhuma nota
        if all(v is None for v in notas):
            stats["sem_nota"] += 1
            continue

        # Série numérica
        serie_raw = SERIE_MAP.get(r["sala_codigo"], 6)
        serie_norm = (serie_raw - SERIE_MIN) / (SERIE_MAX - SERIE_MIN)

        # Médias agregadas
        mg_aluno = sum(medias_aluno.get(r["aluno_id"], [0])) / max(len(medias_aluno.get(r["aluno_id"], [1])), 1)
        mt_vals  = medias_turma.get((r["sala_codigo"], r["materia_id"]), [])
        mt_turma = sum(mt_vals) / len(mt_vals) if mt_vals else 0.0

        ok = ok_aluno.get(r["aluno_id"], [0, 1])
        pct_ok = ok[0] / ok[1] if ok[1] > 0 else 0.0

        # Label
        if mp is None:
            status_encoded = None
            status_label   = None
        elif mp >= 6.0:
            status_encoded = 2
            status_label   = "Aprovado"
            stats["aprovado"] += 1
        elif mp >= 5.0:
            status_encoded = 1
            status_label   = "Recuperação"
            stats["recuperacao"] += 1
        else:
            status_encoded = 0
            status_label   = "Reprovado"
            stats["reprovado"] += 1

        conn.execute("""
            INSERT INTO ml_features (
                aluno_id, materia_id, aluno_nome, materia_nome, sala_nome, serie_num,
                n1, n2, n3, n4, media_ponderada,
                n1_norm, n2_norm, n3_norm, n4_norm, media_pond_norm,
                media_geral_aluno, slope_notas, variancia_notas,
                serie_num_norm, pct_materias_ok, media_turma_norm,
                status_encoded, status_label,
                gerado_em
            ) VALUES (
                ?,?,?,?,?,?,
                ?,?,?,?,?,
                ?,?,?,?,?,
                ?,?,?,
                ?,?,?,
                ?,?,
                datetime('now')
            )
            ON CONFLICT(aluno_id, materia_id) DO UPDATE SET
                aluno_nome        = excluded.aluno_nome,
                materia_nome      = excluded.materia_nome,
                sala_nome         = excluded.sala_nome,
                serie_num         = excluded.serie_num,
                n1=excluded.n1, n2=excluded.n2, n3=excluded.n3, n4=excluded.n4,
                media_ponderada   = excluded.media_ponderada,
                n1_norm           = excluded.n1_norm,
                n2_norm           = excluded.n2_norm,
                n3_norm           = excluded.n3_norm,
                n4_norm           = excluded.n4_norm,
                media_pond_norm   = excluded.media_pond_norm,
                media_geral_aluno = excluded.media_geral_aluno,
                slope_notas       = excluded.slope_notas,
                variancia_notas   = excluded.variancia_notas,
                serie_num_norm    = excluded.serie_num_norm,
                pct_materias_ok   = excluded.pct_materias_ok,
                media_turma_norm  = excluded.media_turma_norm,
                status_encoded    = excluded.status_encoded,
                status_label      = excluded.status_label,
                gerado_em         = datetime('now')
        """, (
            r["aluno_id"], r["materia_id"],
            r["aluno_nome"], r["materia_nome"], r["sala_nome"], serie_raw,
            n1, n2, n3, n4, mp,
            (n1 / 10.0) if n1 is not None else None,
            (n2 / 10.0) if n2 is not None else None,
            (n3 / 10.0) if n3 is not None else None,
            (n4 / 10.0) if n4 is not None else None,
            (mp / 10.0) if mp is not None else None,
            round(mg_aluno / 10.0, 4),
            round(_slope(notas), 4),
            round(_std(notas), 4),
            round(serie_norm, 4),
            round(pct_ok, 4),
            round(mt_turma / 10.0, 4),
            status_encoded, status_label,
        ))
        total += 1

    conn.commit()
    conn.close()
    return total, stats


def exportar_ml_csv(sala_id=None, output_path="ml_dataset.csv"):
    """
    Exporta ml_features como CSV pronto para sklearn / PyTorch / Keras.
    Colunas: 7 features numéricas + label numérico + label texto.
    """
    conn = get_conn()
    q = """
        SELECT
            aluno_nome, materia_nome, sala_nome, serie_num,
            n1_norm, n2_norm, n3_norm, n4_norm,
            media_pond_norm, media_geral_aluno,
            slope_notas, variancia_notas,
            serie_num_norm, pct_materias_ok, media_turma_norm,
            status_encoded, status_label
        FROM ml_features
        WHERE status_encoded IS NOT NULL
    """
    params = []
    if sala_id:
        q += " AND sala_nome = (SELECT nome FROM salas WHERE id=?)"
        params.append(sala_id)
    q += " ORDER BY sala_nome, aluno_nome, materia_nome"
    rows = conn.execute(q, params).fetchall()
    conn.close()

    if not rows:
        return None, "Nenhuma feature gerada. Execute 'Gerar Features ML' primeiro."

    import csv
    headers = [
        "aluno", "materia", "turma", "serie_num",
        # features de entrada (X)
        "n1_norm", "n2_norm", "n3_norm", "n4_norm",
        "media_pond_norm", "media_geral_aluno",
        "slope_notas", "variancia_notas",
        "serie_num_norm", "pct_materias_ok", "media_turma_norm",
        # label (y)
        "status_encoded", "status_label",
    ]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([round(v, 4) if isinstance(v, float) else v for v in r])

    return output_path, f"{len(rows)} amostras exportadas."


def get_ml_stats():
    """Retorna estatísticas da tabela ml_features para exibir na GUI."""
    conn = get_conn()
    total = conn.execute("SELECT COUNT(*) FROM ml_features WHERE status_encoded IS NOT NULL").fetchone()[0]
    dist  = conn.execute("""
        SELECT status_label, COUNT(*) as cnt
        FROM ml_features WHERE status_encoded IS NOT NULL
        GROUP BY status_label
    """).fetchall()
    conn.close()
    return total, {r[0]: r[1] for r in dist}


def exportar_excel(sala_id=None, output_path="notas_exportadas.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    dados = get_relatorio(sala_id)
    if not dados:
        return None, "Nenhum dado com notas para exportar."

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas"

    # Styles
    header_fill = PatternFill("solid", start_color="2D3A8C")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    alt_fill = PatternFill("solid", start_color="EEF1FA")
    reprov_fill = PatternFill("solid", start_color="FFD6D6")
    aprov_fill = PatternFill("solid", start_color="D6FFE4")
    thin = Side(style='thin', color='C0C0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    headers = ["Aluno", "Turma", "Disciplina", "N1", "N2", "N3", "N4", "Média Final", "Status"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 22

    for i, row in enumerate(dados, 2):
        notas = [row.get('n1'), row.get('n2'), row.get('n3'), row.get('n4')]
        notas_validas = [n for n in notas if n is not None]
        media = round(sum(notas_validas) / len(notas_validas), 1) if notas_validas else None
        status = "Aprovado" if media is not None and media >= 6 else "Recuperação" if media is not None else "-"

        values = [
            row['aluno'], row['sala'], row['materia'],
            row.get('n1'), row.get('n2'), row.get('n3'), row.get('n4'),
            media, status
        ]
        ws.append(values)

        fill = alt_fill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        status_fill = aprov_fill if status == "Aprovado" else reprov_fill if status == "Recuperação" else fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col)
            cell.border = border
            cell.alignment = center if col > 2 else Alignment(vertical='center')
            if col == 9:
                cell.fill = status_fill
                cell.font = Font(bold=True, name="Calibri", size=10,
                                 color="1A6B2E" if status == "Aprovado" else "8B0000" if status == "Recuperação" else "000000")
            elif col == 8 and media is not None:
                cell.fill = status_fill
                cell.font = Font(bold=True, name="Calibri", size=10)
                cell.number_format = '0.0'
            elif col in [4, 5, 6, 7] and val is not None:
                cell.fill = fill
                cell.number_format = '0.0'
            else:
                cell.fill = fill

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 10
    ws.column_dimensions['I'].width = 14

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(dados)+1}"

    wb.save(output_path)
    return output_path, f"{len(dados)} registros exportados."

def importar_excel(filepath):
    """
    Importa um arquivo .xlsx para o banco de dados.

    Formatos suportados:
    A) Formato padrão do sistema: colunas Aluno, Turma, Disciplina, N1, N2, N3, N4
    B) Formato simples:           colunas Aluno, Disciplina, P1/N1, P2/N2, P3/N3, P4/N4
       (sem coluna Turma — todos ficam em sala padrão ou indicada)

    Retorna: (total_importados, lista_de_erros, lista_de_avisos)
    """
    import pandas as pd

    erros = []
    avisos = []
    total = 0

    try:
        df = pd.read_excel(filepath, sheet_name=0)
    except Exception as e:
        return 0, [f"Erro ao abrir arquivo: {e}"], []

    # Normalize column names: strip, lower, remove accents
    import unicodedata
    def norm(s):
        s = str(s).strip().lower()
        return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()

    col_map = {norm(c): c for c in df.columns}

    # Detect format
    has_turma = any(k in col_map for k in ["turma", "sala", "classe", "ano"])
    has_aluno = any(k in col_map for k in ["aluno", "nome", "estudante"])
    has_disc  = any(k in col_map for k in ["disciplina", "materia", "matéria", "componente"])

    if not has_aluno:
        return 0, ["Coluna 'Aluno' (ou 'Nome') não encontrada no arquivo."], []
    if not has_disc:
        return 0, ["Coluna 'Disciplina' (ou 'Matéria') não encontrada no arquivo."], []

    # Resolve column aliases
    def get_col(aliases):
        for a in aliases:
            if norm(a) in col_map:
                return col_map[norm(a)]
        return None

    col_aluno = get_col(["Aluno", "Nome", "Estudante"])
    col_disc  = get_col(["Disciplina", "Matéria", "Materia", "Componente"])
    col_turma = get_col(["Turma", "Sala", "Classe", "Ano"]) if has_turma else None
    col_n1    = get_col(["N1", "P1", "Nota1", "Nota 1"])
    col_n2    = get_col(["N2", "P2", "Nota2", "Nota 2"])
    col_n3    = get_col(["N3", "P3", "Nota3", "Nota 3"])
    col_n4    = get_col(["N4", "P4", "Nota4", "Nota 4"])

    conn = get_conn()

    # Build lookup caches
    salas_db   = {s['nome'].lower(): s['id'] for s in get_salas()}
    salas_cod  = {s['codigo'].lower(): s['id'] for s in get_salas()}
    materias_db = {}
    for m in get_materias():
        materias_db[m['nome'].lower()] = m['id']

    def get_or_create_sala(nome_raw):
        if not nome_raw or str(nome_raw).strip() in ("", "nan"):
            # Default to first sala
            first = conn.execute("SELECT id FROM salas ORDER BY id LIMIT 1").fetchone()
            return first['id'] if first else None
        n = str(nome_raw).strip()
        nl = n.lower()
        if nl in salas_db:
            return salas_db[nl]
        if nl in salas_cod:
            return salas_cod[nl]
        # Fuzzy: check if any sala name contains this string
        for k, v in salas_db.items():
            if nl in k or k in nl:
                return v
        # Create new sala
        codigo = n[:4].upper().replace(" ", "")
        try:
            conn.execute("INSERT INTO salas (nome, codigo) VALUES (?,?)", (n, codigo))
            conn.commit()
            sid = conn.execute("SELECT id FROM salas WHERE nome=?", (n,)).fetchone()['id']
            salas_db[nl] = sid
            avisos.append(f"Turma '{n}' não encontrada — criada automaticamente.")
            return sid
        except:
            first = conn.execute("SELECT id FROM salas ORDER BY id LIMIT 1").fetchone()
            return first['id'] if first else None

    def get_or_create_materia(nome_raw):
        if not nome_raw or str(nome_raw).strip() in ("", "nan"):
            return None
        n = str(nome_raw).strip().title()
        nl = n.lower()
        if nl in materias_db:
            return materias_db[nl]
        conn.execute("INSERT OR IGNORE INTO materias (nome) VALUES (?)", (n,))
        conn.commit()
        mid = conn.execute("SELECT id FROM materias WHERE nome=?", (n,)).fetchone()
        if mid:
            materias_db[nl] = mid['id']
            avisos.append(f"Matéria '{n}' criada automaticamente.")
            return mid['id']
        return None

    def get_or_create_aluno(nome_raw, sala_id):
        if not nome_raw or str(nome_raw).strip() in ("", "nan"):
            return None
        n = str(nome_raw).strip().title()
        row = conn.execute(
            "SELECT id FROM alunos WHERE nome=? AND sala_id=?", (n, sala_id)
        ).fetchone()
        if row:
            return row['id']
        # Create
        sala = conn.execute("SELECT codigo FROM salas WHERE id=?", (sala_id,)).fetchone()
        count = conn.execute("SELECT COUNT(*) FROM alunos WHERE sala_id=?", (sala_id,)).fetchone()[0]
        matricula = f"{sala['codigo']}{count+1:04d}"
        try:
            conn.execute(
                "INSERT INTO alunos (nome, sala_id, matricula) VALUES (?,?,?)",
                (n, sala_id, matricula)
            )
            conn.commit()
            return conn.execute(
                "SELECT id FROM alunos WHERE nome=? AND sala_id=?", (n, sala_id)
            ).fetchone()['id']
        except Exception as e:
            erros.append(f"Erro ao criar aluno '{n}': {e}")
            return None

    def parse_nota(val):
        if val is None:
            return None
        try:
            v = float(str(val).strip().replace(",", "."))
            return round(max(0.0, min(10.0, v)), 1)
        except:
            return None

    # Process rows
    df = df.dropna(how="all")
    for idx, row in df.iterrows():
        linha = idx + 2  # Excel row number (1-indexed + header)
        aluno_raw  = row.get(col_aluno, "") if col_aluno else ""
        disc_raw   = row.get(col_disc, "")  if col_disc  else ""
        turma_raw  = row.get(col_turma, "") if col_turma else ""

        if str(aluno_raw).strip() in ("", "nan") and str(disc_raw).strip() in ("", "nan"):
            continue  # skip empty rows silently

        sala_id    = get_or_create_sala(turma_raw)
        materia_id = get_or_create_materia(disc_raw)
        aluno_id   = get_or_create_aluno(aluno_raw, sala_id)

        if not aluno_id:
            erros.append(f"Linha {linha}: nome do aluno inválido ('{aluno_raw}').")
            continue
        if not materia_id:
            erros.append(f"Linha {linha}: disciplina inválida ('{disc_raw}').")
            continue

        n1 = parse_nota(row.get(col_n1)) if col_n1 else None
        n2 = parse_nota(row.get(col_n2)) if col_n2 else None
        n3 = parse_nota(row.get(col_n3)) if col_n3 else None
        n4 = parse_nota(row.get(col_n4)) if col_n4 else None

        try:
            conn.execute("""
                INSERT INTO notas (aluno_id, materia_id, n1, n2, n3, n4)
                VALUES (?,?,?,?,?,?)
                ON CONFLICT(aluno_id, materia_id)
                DO UPDATE SET
                    n1 = COALESCE(excluded.n1, notas.n1),
                    n2 = COALESCE(excluded.n2, notas.n2),
                    n3 = COALESCE(excluded.n3, notas.n3),
                    n4 = COALESCE(excluded.n4, notas.n4)
            """, (aluno_id, materia_id, n1, n2, n3, n4))
            conn.commit()
            total += 1
        except Exception as e:
            erros.append(f"Linha {linha}: erro ao salvar nota — {e}")

    conn.close()
    return total, erros, avisos


if __name__ == "__main__":
    init_db()
    path, msg = exportar_excel()
    print(f"Exportado: {path} - {msg}")